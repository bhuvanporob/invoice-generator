import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from io import BytesIO
from num2words import num2words
import pandas as pd
from openpyxl.styles import Font, Border, Side,Alignment


if "authenticated" not in st.session_state or not st.session_state["authenticated"]:
    st.warning("🚫 You must log in to view this page.")
    st.stop()


st.write(f"Hello, {st.session_state['username']}! You have access to this secure page.")

# Styles
bold_font = Font(name="Arial", bold=True)
italic_font = Font(name="Calibri", italic=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# UI title
st.title("Step 3: Generate & Download Invoice")

# Safety checks
if "invoice_data" not in st.session_state or "products" not in st.session_state or len(st.session_state.products) == 0:
    st.warning("Please complete Step 1 and Step 2 before generating the invoice.")
    st.stop()

products = st.session_state.products
invoice = st.session_state.invoice_data

# Totals and Adjustments
subtotal = sum(p["amount"] for p in products)
st.header("Final Adjustments")

freight = st.number_input("Freight Charges", value=0.0, step=0.1)
round_off = round(subtotal, 0)
discount = st.number_input("Discount on Sales", value=0.0, step=0.1)
final_total = round(freight + round_off - discount, 0)
amount_words = num2words(final_total, to='cardinal', lang='en_IN').upper() + " ONLY"

# Preview
st.subheader("Product List Preview")
df = pd.DataFrame(products)
df.index = range(1, len(df) + 1)
st.dataframe(df[["product_name", "product_code", "hsn", "uom", "qty", "rate", "amount"]].style.format({
    "rate": "₹{:.2f}",
    "amount": "₹{:.2f}"
}))

# Generate
if st.button("Generate Invoice"):
    wb = load_workbook("sample invoice.xlsx")
    ws = wb.active

    # Header replacements
    replacements = {
        "{invoice number}": invoice["invoice_number"],
        "{date}": invoice["invoice_date"].strftime("%d-%m-%Y"),
        "{Buyer name}": invoice["buyer_name"],
        "{Address}": invoice["buyer_address"],
        "{Consignee details}": invoice["consignee_details"],
        "{Consingee Address}": invoice["consignee_address"],
        "{INCO terms}": invoice["inco_terms"],
        "{Payment terms}": invoice["payment_terms"],
        "{PO no}": invoice["po_number"],
        "{PO date}": invoice["po_date"].strftime("%d-%m-%Y"),
        "{final destination}": invoice["final_destination"],
        "{final discharge}": invoice["final_discharge"],
        "{Contact person}": invoice["contact_person"],
        "{emailid}": invoice["email_id"],
        "{Contact details}": invoice["contact_details"],
        "{remarks}": invoice["remarks"],
    }
    # Custom Merging for Addresses and Remarks
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{Address}" in cell.value:
                    start_row = cell.row
                    start_col = cell.column
                    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row + 4, end_column=start_col)

                elif "{Consingee Address}" in cell.value:
                    start_row = cell.row
                    start_col = cell.column
                    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row + 4, end_column=start_col)



    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, value in replacements.items():
                    if key in cell.value:
                        cell.value = cell.value.replace(key, str(value))

    # Wrap text for Buyer Address (update cell ref as per your template)
    ws["F5"].alignment = Alignment(vertical='top',wrap_text=True)

    # Wrap text for Consignee Address
    ws["F13"].alignment = Alignment(vertical='top',wrap_text=True)


    # Insert products
    start_row = 25
    row_pointer = start_row

    for idx, product in enumerate(products):
        ws.insert_rows(row_pointer)
        ws.insert_rows(row_pointer)

        # Main row
        ws.cell(row=row_pointer, column=2, value=idx + 1).font = bold_font
        ws.cell(row=row_pointer, column=2).alignment = Alignment(vertical='center', horizontal='center')
        name_cell = ws.cell(row=row_pointer, column=3, value=product["product_name"])
        name_cell.font = bold_font
        name_cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=row_pointer, column=4, value=product["uom"]).font = bold_font
        ws.cell(row=row_pointer, column=4).alignment = Alignment(vertical='center', horizontal='center')
        ws.cell(row=row_pointer, column=5, value=product["qty"]).font = bold_font
        ws.cell(row=row_pointer, column=5).alignment = Alignment(vertical='center', horizontal='center')
        ws.cell(row=row_pointer, column=6, value=product["rate"]).font = bold_font
        ws.cell(row=row_pointer, column=6).alignment = Alignment(vertical='center', horizontal='center')
        ws.cell(row=row_pointer, column=7, value=product["amount"]).font = bold_font
        ws.cell(row=row_pointer, column=7).alignment = Alignment(vertical='center', horizontal='center')

        # Second row: product code and HSN
        desc = f"Product Code: {product['product_code']}   HSN Code: {product['hsn']}"
        if desc.strip() == "Product Code:    HSN Code:":
            desc = " "

        desc_cell = ws.cell(row=row_pointer + 1, column=3, value=desc)
        desc_cell.font = italic_font
        desc_cell.border = thin_border
        desc_cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Apply border and merge for all columns except col 3
        for col in [2, 4, 5, 6, 7]:
            ws.merge_cells(start_row=row_pointer, start_column=col, end_row=row_pointer+1, end_column=col)
            for r in [row_pointer, row_pointer + 1]:
                ws.cell(row=r, column=col).border = thin_border
                ws.cell(row=r, column=col).alignment = Alignment(vertical='center', horizontal='center')

        # Merge column 3 (description) ONLY for the top row, NOT across
        ws.cell(row=row_pointer, column=3).border = thin_border

        row_pointer += 2

    # Totals
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                label = cell.value.upper()
                if "SUB -TOTAL" in label:
                    ws.cell(row=cell.row, column=cell.column + 1, value=subtotal)
                elif "FREIGHT" in label:
                    ws.cell(row=cell.row, column=cell.column + 1, value=freight)
                elif "ROUND OFF" in label:
                    ws.cell(row=cell.row, column=cell.column + 1, value=round_off)
                elif "DISCOUNT" in label:
                    ws.cell(row=cell.row, column=cell.column + 1, value=discount)
                elif "TOTAL AMOUNT" in label:
                    ws.cell(row=cell.row, column=cell.column + 1, value=final_total)


    # remarks
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "Remarks:" in cell.value:
                start_row = cell.row
                start_col = cell.column + 1
                ws.merge_cells(
                    start_row=start_row,
                    start_column=start_col,
                    end_row=start_row + 2,
                    end_column=start_col + 2
                )
                merged_cell = ws.cell(row=start_row, column=start_col)
                merged_cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)


    # Replace {Amount in words}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{Amount in words}" in cell.value:
                cell.value = cell.value.replace("{Amount in words}", amount_words)
                # Merge 2 cells to the left and 1 below
                start_row = cell.row
                start_col = cell.column 
                end_row = cell.row + 1
                end_col = cell.column + 2
                ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
                merged_cell = ws.cell(row=start_row, column=start_col)
                merged_cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "Bank Details" in cell.value:
                bank_row = cell.row
                bank_col = cell.column + 1  # cell to the right
                ws.merge_cells(
                    start_row=bank_row,
                    start_column=bank_col,
                    end_row=bank_row + 4,
                    end_column=bank_col
                )
                merged_cell = ws.cell(row=bank_row, column=bank_col)
                merged_cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)

    # Save and Download
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("Invoice successfully generated!")

    st.download_button(
        label="Download Invoice Excel",
        data=output,
        file_name=f"Invoice_{invoice['invoice_number']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
